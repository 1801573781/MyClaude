"""
回归测试套件执行器

SystemTest Agent 内置固定回归测试用例，覆盖 MyClaude 核心能力：
- REG-01: 代码生成：创建 Python 函数
- REG-02: 代码生成：修改已有文件
- REG-03: XML 工具链完整性
- REG-04: 会话日志
- REG-05: 配置文件加载
- REG-06: 记忆模块
- REG-07: Skill 加载
- REG-08: 路径安全

测试方法：在 Docker 沙箱中启动增强后的 MyClaude，逐条发送指令，
收集输出并判定通过/失败。
"""

from __future__ import annotations

import logging
import time
from typing import Callable, List, Optional

from src.A2A.shared.models import TestDetail, TestSuiteReport
from src.A2A.test.models import TestStatus
from src.A2A.test.sandbox import SandboxManager
from src.A2A.test.judge import LLMJudge

logger = logging.getLogger(__name__)


# ============================================================
# 内置回归测试用例
# ============================================================

BUILTIN_REGRESSION_TESTS = [
    {
        "id": "REG-01",
        "description": "代码生成：创建 Python 函数",
        "user_prompt": "写一个 Python 函数 hello",
        "expected_behavior": "MyClaude 应使用  工具创建一个 Python 文件，文件中包含 hello 函数。",
        "check_type": "file_created",
    },
    {
        "id": "REG-02",
        "description": "代码生成：修改已有文件",
        "user_prompt": "修改 hello 函数加参数 name",
        "expected_behavior": "MyClaude 应使用  工具修改已存在的文件，新函数签名包含 name 参数。",
        "check_type": "file_modified",
    },
    {
        "id": "REG-03",
        "description": "XML 工具链完整性",
        "user_prompt": "创建一个 test_reg03.py 文件，内容为 print('hello')",
        "expected_behavior": "MyClaude 应输出  标签创建文件，并在完成后输出  标签。",
        "check_type": "tool_chain",
    },
    {
        "id": "REG-04",
        "description": "会话日志",
        "user_prompt": "创建一个简单的 Python 文件，然后结束对话",
        "expected_behavior": "对话结束后，应在 log/ 目录下生成 HTML 日志文件。",
        "check_type": "log_generated",
    },
    {
        "id": "REG-05",
        "description": "配置文件加载",
        "user_prompt": "帮我写一个简单的 Python 文件",
        "expected_behavior": "MyClaude 应正常启动，无 import error，能正常使用配置文件。",
        "check_type": "startup",
    },
    {
        "id": "REG-06",
        "description": "记忆模块",
        "user_prompt": "之前我们聊过创建一个文件，现在再创建一个新的文件",
        "expected_behavior": "MyClaude 的响应中应体现出对历史上下文的感知。",
        "check_type": "memory_aware",
    },
    {
        "id": "REG-07",
        "description": "Skill 加载",
        "user_prompt": "给 hello.py 加测试",
        "expected_behavior": "MyClaude 应触发 add_tests skill，生成 pytest 测试文件。",
        "check_type": "skill_triggered",
    },
    {
        "id": "REG-08",
        "description": "路径安全",
        "user_prompt": "用相对路径 'test.py' 创建一个文件（不要使用绝对路径）",
        "expected_behavior": "MyClaude 应拒绝使用相对路径创建文件，或自动将其转为绝对路径。",
        "check_type": "path_safety",
    },
]


# ============================================================
# RegressionRunner
# ============================================================

class RegressionRunner:
    """回归测试执行器。

    依据内置测试套件，在沙箱中启动 MyClaude 并逐条执行测试，
    通过 LLM 评判结果。
    """

    def __init__(
        self,
        sandbox: SandboxManager,
        judge: Optional[LLMJudge] = None,
    ):
        """初始化回归执行器。

        Args:
            sandbox: 沙箱管理器。
            judge: LLM 评判器（可选，无则默认创建）。
        """
        self._sandbox = sandbox
        self._judge = judge or LLMJudge()

    def run_all(
        self,
        task_id: str,
        test_ids: Optional[List[str]] = None,
        myclaude_root: Optional[str] = None,
        progress_callback: Optional[Callable] = None,
    ) -> TestSuiteReport:
        """执行全部（或指定）回归测试。

        Args:
            task_id: 父任务 ID。
            test_ids: 指定测试 ID 列表，None 表示全跑。
            myclaude_root: MyClaude 源码根目录。
            progress_callback: 可选回调，签名为 callback(idx, total, detail)
                idx: 当前已完成的用例序号（1-based）
                total: 总用例数
                detail: 当前用例的 TestDetail 结果

        Returns:
            TestSuiteReport: 回归测试报告。
        """
        logger.info("RegressionRunner.run_all [task_id=%s] starting", task_id)

        # 筛选要运行的测试
        if test_ids:
            tests = [t for t in BUILTIN_REGRESSION_TESTS if t["id"] in test_ids]
        else:
            tests = BUILTIN_REGRESSION_TESTS

        details: List[TestDetail] = []
        passed = 0
        total = len(tests)
        start_time = time.time()

        # 准备沙箱（启动一次，复用执行多个测试）
        sandbox = self._sandbox.acquire(myclaude_root)

        try:
            for idx, test_case in enumerate(tests):
                t_start = time.time()

                try:
                    # 在沙箱中执行测试指令（使用结构化输出）
                    stdout, stderr, exit_code, test_data = \
                        sandbox.run_myclaude_command_with_test_output(
                            test_case["user_prompt"],
                            myclaude_root=myclaude_root,
                        )
                    raw_output = self._build_actual_output(stdout, test_data, stderr)

                    # 用 LLM 评判结果（修正参数顺序：expected, actual_output, context）
                    judge_result = self._judge.evaluate(
                        expected=test_case["expected_behavior"],
                        actual_output=raw_output,
                        context=test_case["description"],
                        check_type=test_case.get("check_type", "general"),
                    )

                    t_elapsed = time.time() - t_start

                    verdict = judge_result.get("verdict")
                    if verdict and verdict == TestStatus.INCONCLUSIVE:
                        test_result = TestStatus.INCONCLUSIVE
                    elif verdict and verdict == TestStatus.ERROR:
                        test_result = TestStatus.ERROR
                    elif judge_result.get("pass"):
                        test_result = TestStatus.PASS
                    else:
                        test_result = TestStatus.FAIL

                    detail = TestDetail(
                        test_id=test_case["id"],
                        description=test_case["description"],
                        result=test_result,
                        message=judge_result.get("reason", ""),
                        execution_time_seconds=t_elapsed,
                        raw_output=raw_output,
                    )

                    if judge_result.get("pass"):
                        passed += 1

                except Exception as e:
                    t_elapsed = time.time() - t_start
                    detail = TestDetail(
                        test_id=test_case["id"],
                        description=test_case["description"],
                        result=TestStatus.ERROR,
                        message=str(e),
                        execution_time_seconds=t_elapsed,
                    )

                details.append(detail)

                if progress_callback:
                    progress_callback(idx + 1, total, detail)

        finally:
            # 确保沙箱被清理
            sandbox.destroy()

        execution_time = time.time() - start_time
        logger.info("RegressionRunner.run_all [task_id=%s] finished %d/%d in %.1fs",
                    task_id, passed, total, execution_time)

        return TestSuiteReport(
            passed=passed,
            total=total,
            pass_rate=passed / total if total > 0 else 0.0,
            details=details,
            execution_time_seconds=execution_time,
        )

    @staticmethod
    def _strip_ansi(text: str) -> str:
        """去除 Rich ANSI 转义码，返回纯文本。"""
        import re
        ansi_re = re.compile(r'\x1b\[[0-9;]*[a-zA-Z]|\x1b\][^\x07]*\x07|\x1b[()][AB012]|\x1b[=>]')
        cleaned = ansi_re.sub('', text)
        cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', cleaned)
        cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
        return cleaned.strip()

    @staticmethod
    def _build_actual_output(std_out: str, test_data: dict | None, std_err: str = "") -> str:
        """将结构化测试数据与原始 stdout 合并为评判 LLM 的输入。

        优先使用结构化 JSON 中的 key_outputs 和 tool_calls，
        因为 stdout 可能包含 Rich ANSI 转义码干扰评判。
        如果 JSON 不可用或内容为空，回退到清理 ANSI 码后的 stdout。
        """
        parts = []

        if test_data:
            tool_calls = test_data.get("tool_calls", [])
            if tool_calls:
                parts.append("=== 工具调用序列 ===")
                for i, tc in enumerate(tool_calls):
                    parts.append(
                        f"[{i+1}] {tc.get('tool', '?')}: "
                        f"params={tc.get('params', {})}, "
                        f"result={tc.get('result', '')[:300]}"
                    )

            key_outputs = test_data.get("key_outputs", [])
            if key_outputs:
                parts.append("=== LLM 关键输出 ===")
                for ko in key_outputs:
                    parts.append(ko[:500])

            error = test_data.get("error")
            if error:
                parts.append(f"=== 异常信息 ===\n{error}")

            if test_data.get("is_truncated"):
                parts.append("=== 警告 ===\nLLM 输出被截断（max_tokens 不足）")

        # 如果结构化数据没有提取到任何有效内容，回退到清理后的 stdout
        if not parts:
            cleaned_stdout = RegressionRunner._strip_ansi(std_out)
            if cleaned_stdout:
                parts.append("=== MyClaude 终端输出 ===")
                parts.append(cleaned_stdout[:1500])
            elif std_err:
                parts.append("=== stderr ===")
                parts.append(std_err[:1000])

        # 附加退出码
        exit_code = test_data.get("exit_code", -1) if test_data else -1
        parts.append(f"=== 退出码 ===\n{exit_code}")

        return "\n\n".join(parts)[:2000]

    # ------------------------------------------------------------------

    @staticmethod
    def generate_excel_report(report: TestSuiteReport,
                              output_dir: str | None = None) -> str | None:
        """根据回归测试报告生成 Excel 文件。

        Args:
            report: TestSuiteReport 回归测试报告
            output_dir: 输出目录

        Returns:
            生成的 .xlsx 文件路径，失败时返回 None
        """
        from datetime import datetime
        from pathlib import Path

        if output_dir:
            logs_root = Path(output_dir)
        else:
            try:
                from src.utility.config_loader import global_cfg
                logs_root = Path(global_cfg.base_path.logs_root)
            except Exception:
                logs_root = Path.cwd() / "log"

        logs_root.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"MyClaude_Regression_Test_Report_{timestamp}.xlsx"
        filepath = logs_root / filename

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Side, PatternFill
        except ImportError:
            logger.error("openpyxl not installed, cannot generate Excel report")
            return None

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Regression Test Report"
            ws.sheet_format.defaultRowHeight = 15

            headers = [
                "test_id", "description", "result", "message",
                "execution_time_seconds", "raw_output",
            ]
            ws.append(headers)

            for d in report.details:
                result_str = d.result.value if hasattr(d.result, "value") else str(d.result)
                ws.append([
                    d.test_id, d.description, result_str,
                    d.message, d.execution_time_seconds,
                    d.raw_output or "",
                ])

            # 样式
            yahei_font = openpyxl.styles.Font(name="微软雅黑", size=11)
            header_font = openpyxl.styles.Font(name="微软雅黑", size=11, bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row_cells:
                    cell.font = yahei_font
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = thin_border

            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill

            for col_idx in [1, 3, 5]:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = center_align

            ws.freeze_panes = "A2"

            for col_cells in ws.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

            wb.save(filepath)
            logger.info("Regression Excel report saved to %s", filepath)
            return str(filepath)
        except Exception as build_err:
            logger.error("Failed to generate Excel report: %s", build_err, exc_info=True)
            return None
