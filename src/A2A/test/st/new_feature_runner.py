"""
新功能测试执行器

针对 MyClaude 新增能力，逐个执行验收用例，调用评判 LLM 判定 PASS / FAIL。
"""

from __future__ import annotations

import logging
import time
from typing import Callable, Optional

from ..models import TestCase, TestResult, TestStatus
from ..sandbox import SandboxManager
from ..judge import LLMJudge

logger = logging.getLogger(__name__)


class NewFeatureRunner:
    """新功能测试用例执行器"""

    def __init__(self, sandbox_mgr: SandboxManager, judge: LLMJudge):
        self._sandbox_mgr = sandbox_mgr
        self._judge = judge

    # ------------------------------------------------------------------

    def execute(self,
                test_cases: list,
                myclaude_root: str | None = None,
                progress_callback: Optional[Callable] = None) -> list[TestResult]:
        """执行全部新功能测试用例，返回结果列表

        Args:
            test_cases: 测试用例列表（支持 TestCase 对象或 dict）
            myclaude_root: MyClaude 源码根目录
            progress_callback: 可选回调，签名为 callback(idx, total, result)
                idx: 当前已完成的用例序号（1-based）
                total: 总用例数
                result: 当前用例的 TestResult 结果
        """
        results: list[TestResult] = []
        total = len(test_cases)

        for i, raw_case in enumerate(test_cases):
            # 归一化：dict → TestCase，统一用属性访问
            if isinstance(raw_case, dict):
                case = TestCase(**raw_case)
            else:
                case = raw_case
            logger.info("Running new-feature case [id=%s] %s", case.id, case.description)
            result = self._run_one(case, myclaude_root)
            results.append(result)
            logger.info("Case [id=%s] -> %s", case.id, result.status)

            if progress_callback:
                progress_callback(i + 1, total, result)

        return results

    # ------------------------------------------------------------------

    def _run_one(self, case, myclaude_root: str | None) -> TestResult:
        t0 = time.perf_counter()
        sandbox = None

        try:
            # 1. 在沙箱中启动 MyClaude 并发送指令，获取结构化测试结果
            sandbox = self._sandbox_mgr.acquire()
            std_out, std_err, exit_code, test_data = sandbox.run_myclaude_command_with_test_output(
                user_prompt=case.user_prompt,
                myclaude_root=myclaude_root,
            )

            # 2. 构建评判 actual_output（优先使用结构化 JSON 的关键输出片段）
            actual_output = self._build_actual_output(std_out, test_data)

            # 3. 确定 check_type
            check_type = getattr(case, 'check_type', None) or "general"

            # 4. 调用评判 LLM
            if exit_code == 0:
                verdict_result = self._judge.evaluate(
                    expected=case.expected_behavior,
                    actual_output=actual_output,
                    context=case.description,
                    check_type=check_type,
                )
                judge_verdict = verdict_result.get("verdict")
                if judge_verdict and judge_verdict in (
                    TestStatus.PASS, TestStatus.FAIL,
                    TestStatus.INCONCLUSIVE, TestStatus.ERROR,
                ):
                    verdict = judge_verdict
                elif verdict_result.get("pass"):
                    verdict = TestStatus.PASS
                else:
                    verdict = TestStatus.FAIL
            else:
                verdict = TestStatus.FAIL

            elapsed = round(time.perf_counter() - t0, 2)
            return TestResult(
                test_id=case.id,
                description=case.description,
                status=verdict,
                stdout_preview=actual_output[:500] if actual_output else "(empty)",
                stderr_preview=std_err[:500] if std_err else "",
                exit_code=exit_code,
                duration_seconds=elapsed,
            )

        except Exception as exc:
            elapsed = round(time.perf_counter() - t0, 2)
            logger.exception("New-feature case [id=%s] crashed", case.id)
            return TestResult(
                test_id=case.id,
                description=case.description,
                status=TestStatus.ERROR,
                stdout_preview=str(exc)[:500],
                stderr_preview="",
                exit_code=-1,
                duration_seconds=elapsed,
            )

        finally:
            # 只在成功 acquire 后才 release，防止操作未设置的容器
            if sandbox is not None:
                self._sandbox_mgr.release(sandbox)

    # ------------------------------------------------------------------

    @staticmethod
    def _build_actual_output(std_out: str, test_data: dict | None) -> str:
        """将结构化测试数据与原始 stdout 合并为评判 LLM 的输入。

        优先使用结构化 JSON 中的 key_outputs（LLM 纯文本片段）和 tool_calls
        （工具调用记录），因为 stdout 可能包含 Rich ANSI 转义码干扰评判。
        如果 JSON 不可用，回退到原始 stdout。
        """
        if not test_data:
            return std_out[:2000]

        parts = []

        # 1. 工具调用摘要（结构化，无 Rich 转义码）
        tool_calls = test_data.get("tool_calls", [])
        if tool_calls:
            parts.append("=== 工具调用序列 ===")
            for i, tc in enumerate(tool_calls):
                parts.append(
                    f"[{i+1}] {tc.get('tool', '?')}: "
                    f"params={tc.get('params', {})}, "
                    f"result={tc.get('result', '')[:300]}"
                )

        # 2. LLM 输出的纯文本片段
        key_outputs = test_data.get("key_outputs", [])
        if key_outputs:
            parts.append("=== LLM 关键输出 ===")
            for ko in key_outputs:
                parts.append(ko[:500])

        # 3. 错误信息
        error = test_data.get("error")
        if error:
            parts.append(f"=== 异常信息 ===\n{error}")

        # 4. 退出码
        parts.append(f"=== 退出码 ===\n{test_data.get('exit_code', -1)}")

        # 5. 截断标记
        if test_data.get("is_truncated"):
            parts.append("=== 警告 ===\nLLM 输出被截断（max_tokens 不足）")

        return "\n\n".join(parts)[:2000]

    # ------------------------------------------------------------------

    @staticmethod
    def generate_excel_report(results: list[TestResult],
                              output_dir: str | None = None) -> str | None:
        """根据新功能测试结果生成 Excel 文件。

        Args:
            results: TestResult 列表
            output_dir: 输出目录

        Returns:
            生成的 .xlsx 文件路径，失败时返回 None
        """
        from datetime import datetime
        from pathlib import Path

        if output_dir:
            logs_root = Path(output_dir)
        else:
            try:
                from src.utility.config_loader import global_cfg
                logs_root = Path(global_cfg.base_path.logs_root)
            except Exception:
                logs_root = Path.cwd() / "log"

        logs_root.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"MyClaude_NewFeature_Test_Report_{timestamp}.xlsx"
        filepath = logs_root / filename

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Side, PatternFill
        except ImportError:
            logger.error("openpyxl not installed, cannot generate Excel report")
            return None

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "NewFeature Test Report"
            ws.sheet_format.defaultRowHeight = 15

            headers = [
                "test_id", "description", "status", "exit_code",
                "duration_seconds", "stdout_preview", "stderr_preview",
            ]
            ws.append(headers)

            for r in results:
                status_str = r.status.value if hasattr(r.status, "value") else str(r.status)
                ws.append([
                    r.test_id, r.description, status_str, r.exit_code,
                    r.duration_seconds, r.stdout_preview, r.stderr_preview,
                ])

            # 样式
            yahei_font = openpyxl.styles.Font(name="微软雅黑", size=11)
            header_font = openpyxl.styles.Font(name="微软雅黑", size=11, bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row_cells:
                    cell.font = yahei_font
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = thin_border

            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill

            for col_idx in [1, 3, 4, 5]:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = center_align

            ws.freeze_panes = "A2"

            for col_cells in ws.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

            wb.save(filepath)
            logger.info("NewFeature Excel report saved to %s", filepath)
            return str(filepath)
        except Exception as build_err:
            logger.error("Failed to generate Excel report: %s", build_err, exc_info=True)
            return None
