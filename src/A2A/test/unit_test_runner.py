"""
单元测试执行器

直接 import 被测模块、调用被测函数、捕获返回值/异常，再调用 LLMJudge 评判结果。
复用现有的 LLMJudge 评判逻辑，不做端到端 CLI 调用。
"""

from __future__ import annotations

import importlib
import logging
import time
import traceback
from pathlib import Path

from src.A2A.test.models import UnitTestResult, TestStatus
from src.A2A.test.judge import LLMJudge

logger = logging.getLogger(__name__)


class UnitTestRunner:
    """单元测试用例执行器"""


    def __init__(self, judge: LLMJudge):
        self._judge = judge

    # ------------------------------------------------------------------

    def execute(self,
                test_cases: list,
                myclaude_root: str | None = None,
                progress_callback: callable | None = None) -> list[UnitTestResult]:
        """执行全部单元测试用例，返回结果列表

        Args:
            progress_callback: 可选回调，签名为 callback(idx, total, results)
                idx: 当前已完成的用例序号（1-based）
                total: 总用例数
                results: 已完成用例的结果列表
        """
        results: list[UnitTestResult] = []
        total = len(test_cases)

        for i, case in enumerate(test_cases):
            logger.info("Running unit-test case [id=%s] %s", case["id"], case["description"])
            result = self._run_one(case, myclaude_root)
            # 注入原始用例数据，供 Excel 报告使用
            result._case = case
            results.append(result)
            logger.info("Case [id=%s] -> %s", case["id"], result.status)
            logger.info("\n\n")

            if progress_callback:
                progress_callback(i + 1, total, results)

        return results

    # ------------------------------------------------------------------

    def _run_one(self, case, myclaude_root: str | None) -> UnitTestResult:
        t0 = time.perf_counter()

        try:
            # 1. 动态导入被测模块并调用函数
            actual_output = self._invoke_target(
                target_module=case["target_module"],
                target_function=case["target_function"],
                test_input=case.get("test_input", ""),
                param_types=case.get("param_types", {}),
                myclaude_root=myclaude_root,
            )

            # 2. 调用评判 LLM

            verdict_result = self._judge.evaluate(
                expected=case["expected_behavior"],
                actual_output=actual_output,
                context=case["description"],
                check_type=case.get("check_type", "general"),
            )

            verdict = verdict_result.get("verdict")
            if verdict and verdict in (TestStatus.PASS, TestStatus.FAIL, TestStatus.INCONCLUSIVE, TestStatus.ERROR):
                status = verdict
            else:
                status = TestStatus.PASS if verdict_result.get("pass") else TestStatus.FAIL
            elapsed = round(time.perf_counter() - t0, 2)
            reason = verdict_result.get("reason", "") or "（评判 LLM 未返回理由）"

            return UnitTestResult(
                test_id=case["id"],
                description=case["description"],
                status=status,
                actual_output=actual_output[:500],
                reason=reason,
                duration_seconds=elapsed,
            )

        except Exception as exc:
            elapsed = round(time.perf_counter() - t0, 2)
            logger.exception("Unit-test case [id=%s] crashed", case["id"])
            return UnitTestResult(
                test_id=case["id"],
                description=case["description"],
                status=TestStatus.ERROR,
                actual_output=traceback.format_exc(),
                reason=str(exc),
                duration_seconds=elapsed,
            )

    # ------------------------------------------------------------------

    @staticmethod
    def generate_excel_report(results: list[UnitTestResult],
                              myclaude_root: str | None = None,
                              output_dir: str | None = None) -> Path:
        """根据测试结果生成 Excel 报告，输出到 logs_root 目录。

        Args:
            results: UnitTestResult 列表（每个元素需携带 _case 原始用例数据）
            myclaude_root: MyClaude 源码根目录（用于定位 config，output_dir 提供时优先使用）
            output_dir: 输出目录（优先使用；未提供时从 config 读取 logs_root）

        Returns:
            生成的 .xlsx 文件路径
        """
        import sys
        from datetime import datetime
        from pathlib import Path

        if output_dir:
            logs_root = Path(output_dir)
        else:
            root = myclaude_root or str(Path(__file__).resolve().parents[3])
            if root not in sys.path:
                sys.path.insert(0, root)
            from utility.config_loader import global_cfg
            logs_root = Path(global_cfg.base_path.logs_root)

        logs_root.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"MyClaude_Unit_Test_Report_{timestamp}.xlsx"
        filepath = logs_root / filename

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Side, PatternFill
        except ImportError:
            logger.error("openpyxl not installed, cannot generate Excel report")
            return filepath

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unit Test Report"

        # 设置默认行高为 15，使内容更易阅读
        ws.sheet_format.defaultRowHeight = 15

        # 表头：用例原始列 + 测试结果列
        headers = [
            "id", "description", "target_module", "target_function",
            "test_input", "expected_behavior",
            "status", "actual_output", "reason", "duration_seconds",
        ]
        ws.append(headers)

        for result in results:
            case = getattr(result, "_case", {})
            row = [
                case.get("id", result.test_id),
                case.get("description", result.description),
                case.get("target_module", ""),
                case.get("target_function", ""),
                case.get("test_input", ""),
                case.get("expected_behavior", ""),
                result.status.value if hasattr(result.status, "value") else str(result.status),
                result.actual_output,
                result.reason,
                result.duration_seconds,
            ]
            ws.append(row)

        # --- 样式定义 ---
        yahei_font = openpyxl.styles.Font(name="微软雅黑", size=11)
        header_font = openpyxl.styles.Font(name="微软雅黑", size=11, bold=True)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # 浅灰
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # --- 应用全局样式：所有单元格上下居中 + 自动换行 + 微软雅黑 + 四面框线 ---
        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row_cells:
                cell.font = yahei_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = thin_border

        # --- 首行样式：居中、加粗、浅灰底色 ---
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = header_fill

        # --- A列(1)、G列(7)、J列(10) 左右居中 ---
        center_columns = [1, 7, 10]
        for col_idx in center_columns:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.alignment = center_align

        # --- 冻结首行 ---
        ws.freeze_panes = "A2"

        # --- 自动调整列宽 ---
        for col_cells in ws.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

        wb.save(filepath)
        logger.info("Excel report saved to %s", filepath)
        return filepath

    # ------------------------------------------------------------------

    @staticmethod
    def _invoke_target(target_module: str,
                       target_function: str,
                       test_input: str,
                       param_types: dict | None = None,
                       myclaude_root: str | None = None) -> str:
        """动态导入被测模块并调用函数，返回 repr(result) 或异常字符串。

        Args:
            target_module: 如 'src.utility.file_tool'
            target_function: 如 'resolve_path'
            test_input: 测试说明字符串，复杂参数需 Runner 内部构造（见 _build_args）
            param_types: 参数名→类型提示的映射字典，用于类型强制转换
            myclaude_root: MyClaude 源码根目录

        Returns:
            格式化的实际输出字符串（含返回值或异常信息）
        """
        # 确保项目根在 sys.path 中
        import sys
        root = myclaude_root or str(Path(__file__).resolve().parents[3])
        if root not in sys.path:
            sys.path.insert(0, root)

        # 动态导入模块
        mod = importlib.import_module(target_module)
        func = getattr(mod, target_function)

        # 构造参数（根据 test_input 解析）
        args, kwargs = UnitTestRunner._build_args(
            test_input, target_function, param_types or {}
        )

        # 调用函数
        result = func(*args, **kwargs)

        # 格式化返回值
        return f"返回值: {repr(result)}"

    # ------------------------------------------------------------------

    @staticmethod
    def _build_args(test_input: str,
                    target_function: str,
                    param_types: dict | None = None) -> tuple[list, dict]:
        """根据 test_input 描述构造函数参数。

        优先尝试统一的键值对格式：'key1' : 'value1', 'key2' : 'value2'
        键名对应被测函数的参数名。解析成功后全部以关键字参数形式传递。
        根据 param_types 进行类型强制转换（int/float/bool/Path/NoneType）。
        如果无法解析为键值对，回退到函数特定的旧格式解析逻辑。
        """
        import re

        if param_types is None:
            param_types = {}

        # ── 类型强制转换辅助函数 ──
        def _coerce_value(key: str, val: str):
            """根据 param_types 将字符串值转为对应 Python 类型。"""
            hint = param_types.get(key, "")
            hint_lower = hint.lower().strip()
            # 剥离 Optional/Union 包装
            hint_lower = re.sub(r"^optional\[(.*)\]$", r"\1", hint_lower, flags=re.IGNORECASE)
            hint_lower = re.sub(r"^union\[(.*)\]$", r"\1", hint_lower, flags=re.IGNORECASE)

            # 如果 hint 包含 None 且值为 None → 返回 None
            if "none" in hint_lower.split(","):
                if val.strip().lower() == "none":
                    return None

            base_types = [t.strip().strip("[]") for t in hint_lower.split(",")]

            for bt in base_types:
                bt_lower = bt.lower()
                if bt_lower in ("int", "integer"):
                    try:
                        return int(float(val))
                    except (ValueError, TypeError):
                        pass
                elif bt_lower in ("float", "number"):
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        pass
                elif bt_lower in ("bool", "boolean"):
                    v_lower = val.strip().lower()
                    if v_lower in ("true", "1"):
                        return True
                    elif v_lower in ("false", "0"):
                        return False
                    # 非标准值，保持原值
                elif bt_lower in ("str", "string"):
                    return str(val)
                elif bt_lower in ("path", "pathlike", "purepath"):
                    from pathlib import Path
                    return Path(val) if val else None
                # 列表、字典等保持字符串，由函数本身处理
            return val  # 无法匹配任何类型，保留原字符串

        # ── 特殊处理：参数结构复杂的函数（不适合键值对格式） ──

        # execute_code_tool: test_input 描述工具调用，需构造复合 tool_dict
        if target_function == "execute_code_tool":
            tool_match = re.search(r'(file_view|create|str_replace|bash|done|use_skill)', test_input)
            tool_name = tool_match.group(1) if tool_match else "file_view"

            tool_params = {}

            path_match = re.search(r"path=['\"]([^'\"]+)['\"]", test_input)
            if path_match:
                tool_params["path"] = path_match.group(1)
            limit_match = re.search(r"limit=['\"]?(\d+)", test_input)
            if limit_match:
                tool_params["limit"] = int(limit_match.group(1))
            offset_match = re.search(r"offset=['\"]?(\d+)", test_input)
            if offset_match:
                tool_params["offset"] = int(offset_match.group(1))
            summary_match = re.search(r"summary=['\"]([^'\"]*)['\"]", test_input)
            if summary_match:
                tool_params["summary"] = summary_match.group(1)

            if tool_name == "create":
                body_match = re.search(
                    r"(?:body|content)=['\"]([^'\"]+)['\"]", test_input
                )
                tool_params["content"] = body_match.group(1) if body_match else ""

            if tool_name == "str_replace":
                old_match = re.search(r"old=['\"]([^'\"]*)['\"]", test_input)
                new_match = re.search(r"new=['\"]([^'\"]*)['\"]", test_input)
                if old_match:
                    tool_params["old"] = old_match.group(1)
                if new_match:
                    tool_params["new"] = new_match.group(1)

            if tool_name == "use_skill":
                name_match = re.search(r"name=['\"]([^'\"]+)['\"]", test_input)
                if name_match:
                    tool_params["name"] = name_match.group(1)

            tool_dict = {"llm_tool": tool_name, "params": tool_params}
            return [tool_dict], {}

        # append_tool_exec_result: 参数是嵌套列表/字典，不适合键值对
        if target_function == "append_tool_exec_result":
            return [], {"api_messages": [
                {"role": "system", "content": "初始系统提示词"},
                {"role": "user", "content": "用户输入"},
            ], "tool_exec_result": {"role": "user", "content": "工具执行结果"}}

        # ── 通用键值对解析：'key' : 'value', ... ──
        # 值部分允许空字符串（如 's' : '' 或 'name' : ""）
        kv_pattern = r"""['"]([^'"]+)['"]\s*:\s*['"]([^'"]*)['"]"""
        kv_matches = re.findall(kv_pattern, test_input)

        if kv_matches:
            kwargs = {}
            for k, v in kv_matches:
                kwargs[k] = _coerce_value(k, v)

            return [], kwargs

        # ── 回退：旧格式的函数特定解析 ──

        # resolve_path: test_input 形如 "绝对路径 'D:/...' 和相对路径 'test.py'"
        if target_function == "resolve_path":
            paths = re.findall(r"['\"]([^'\"]+)['\"]", test_input)
            return paths, {}

        # parse_tools: test_input 是主 content
        if target_function == "parse_tools":
            return [test_input], {}

        # file_create: 旧自然语言格式回退
        if target_function == "file_create":
            paths = re.findall(r"([A-Za-z]:/\S+\.py)", test_input)
            if len(paths) >= 2:
                return [paths[0], "print('v2')"], {}
            return [test_input], {}

        # strip_thinking: test_input 是待处理的文本
        if target_function == "strip_thinking":
            return [test_input], {}

        # 默认：test_input 作为单参数字符串传入
        return [test_input], {}


if __name__ == "__main__":
    # 配置日志输出到控制台和文件
    from src.utility.config_loader import global_cfg

    logs_root = Path(global_cfg.base_path.logs_root)
    logs_root.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[
            logging.FileHandler(
                logs_root / "unit_test_runner.log",
                encoding="utf-8",
            ),
        ],
    )

    judege = LLMJudge()
    ut = UnitTestRunner(judge=judege)

    ut_test_cases = [
        {
            "id": "UT-FO-001",
            "description": "file_create 在文件不存在时成功创建新文件",
            "target_module": "src.utility.file_tool",
            "target_function": "file_create",
            "test_input": "'root' : 'D:/AI/MyClaude/code_output', 'path' : 'test_fo001.py', 'content' : 'x = 1'",
            "expected_behavior": "file_create 应在 D:/AI/MyClaude/code_output/ 目录下创建 test_fo001.py 文件，文件内容为 'x = 1'。返回结果不应包含 [BLOCKED] 或 [ERROR]，应包含成功创建的信息。",
            "check_type": "file_created"
        },
    ]

    myclaude_root_path = global_cfg.base_path.project_root

    results = ut.execute(test_cases=ut_test_cases, myclaude_root=myclaude_root_path)

    # 生成 Excel 报告，输出到 logs_root
    report_path = UnitTestRunner.generate_excel_report(
        results, output_dir=str(logs_root)
    )
    print(f"Excel report saved to: {report_path}")
